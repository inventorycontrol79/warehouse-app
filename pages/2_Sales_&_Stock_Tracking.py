import streamlit as st
import pandas as pd
import numpy as np
import json
import gspread
import io  
from google.oauth2.service_account import Credentials
from copilot import inject_floating_copilot
from datetime import datetime, timedelta

st.set_page_config(page_title="SABIN PLASTIC // Inventory Intelligence", layout="wide")

if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

if "df_stock_live" not in st.session_state:
    st.session_state.df_stock_live = None

url_params = st.query_params
if url_params.get("key", "") == "sabin_inventory":
    st.session_state.is_admin = True

is_admin = st.session_state.is_admin

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght=300;400;600;800&display=swap');
    .stApp { background-color: #0B0F19; color: #E2E8F0; font-family: 'Plus Jakarta Sans', sans-serif; }
    h1, h2, h3, h4, h5, h6, [data-testid="stMarkdownContainer"] p { color: #F8FAFC !important; }
    label, .stWidgetLabel p { color: #94A3B8 !important; font-weight: 600 !important; }
    .premium-header { border-bottom: 1px solid #1E293B; padding-bottom: 1.5rem; margin-bottom: 2rem; margin-top: 1rem; }
    .sabin-logo { font-size: 32px; font-weight: 800; letter-spacing: 4px; color: #F8FAFC !important; margin: 0; line-height: 1.2; }
    .sabin-logo span { color: #0EA5E9 !important; }
    .sabin-sub { font-size: 12px; font-weight: 600; letter-spacing: 3px; color: #94A3B8 !important; text-transform: uppercase; margin-top: 4px; }
    section[data-testid="stSidebar"] { background-color: #0F172A !important; border-right: 1px solid #1E293B; }
    div[data-testid="metric-container"] { background-color: #111827; border: 1px solid #1E293B; border-top: 3px solid #0EA5E9; border-radius: 6px; padding: 20px; }
    .upload-box { background-color: #111827; border: 1px dashed #1E293B; border-radius: 8px; padding: 20px; margin-bottom: 20px; }
    .admin-box { background-color: #1E1B4B; border: 1px solid #4338CA; border-radius: 8px; padding: 20px; margin-top: 20px; }
    
    div[data-testid="stRadio"] > label { display: none; }
    div[data-testid="stRadio"] div[role="radiogroup"] { gap: 10px; }
    div[data-testid="stRadio"] label[data-baseweb="radio"] {
        background-color: #111827;
        border: 1px solid #1E293B;
        padding: 8px 16px;
        border-radius: 20px;
        color: #94A3B8;
        transition: all 0.2s ease-in-out;
    }
    div[data-testid="stRadio"] label[data-baseweb="radio"]:hover {
        border-color: #0EA5E9;
        color: #F8FAFC;
    }
    div[data-testid="stRadio"] label[data-baseweb="radio"][data-checked="true"] {
        background-color: #0EA5E9 !important;
        border-color: #0EA5E9 !important;
        color: #0B0F19 !important;
        font-weight: bold;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("<div class='premium-header'><div class='sabin-logo'>SABIN <span>PLASTIC</span></div><div class='sabin-sub'>Enterprise Warehouse Tracking System</div></div>", unsafe_allow_html=True)

def clean_item_code(val):
    val_str = str(val).strip()
    if val_str.endswith(".0"): return val_str[:-2]
    return val_str

def get_google_client():
    try:
        raw_json = st.secrets["GCP_JSON"]
        creds_dict = json.loads(raw_json)
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as e:
        st.error(f"🚨 Authentication Failed: {e}")
        return None

@st.cache_resource(ttl=3600)
def get_google_sheet_file():
    gc = get_google_client()
    if not gc: return None
    try: return gc.open_by_url(st.secrets["GSHEET_URL"])
    except Exception as e:
        st.error(f"🚨 Sheet Connection Failed: {e}")
        return None

def get_fresh_google_sheet_file():
    gc = get_google_client()
    if not gc: return None
    try: return gc.open_by_url(st.secrets["GSHEET_URL"])
    except Exception: return None

@st.cache_data(ttl=60) 
def load_all_inventory_data():
    fallback_data = {3: [], 4: [], 5: []}
    sh = get_google_sheet_file()
    if not sh: return fallback_data
        
    def safe_get_records(ws):
        try:
            raw_data = ws.get_all_values()
            if not raw_data: return []
            headers = [str(h).strip() for h in raw_data[0]]
            rows = raw_data[1:]
            df = pd.DataFrame(rows, columns=headers)
            if "" in df.columns: df = df.drop(columns=[""])
            return df.to_dict(orient="records")
        except Exception:
            return []

    try:
        ws3_data = safe_get_records(sh.get_worksheet(3))
        try: ws4_target = sh.worksheet("Daily_Snapshot_Log")
        except Exception: ws4_target = sh.get_worksheet(4)
        ws4_data = safe_get_records(ws4_target)
        ws5_data = safe_get_records(sh.get_worksheet(5))
        return {3: ws3_data, 4: ws4_data, 5: ws5_data}
    except Exception as e:
        st.error(f"🚨 Google Sheets Quota Error: {e}")
        return fallback_data

sheet_payload = load_all_inventory_data()

TARGET_STOCK_COLS = [
    "Item_Code", "Item_Name", "Product_Category", "Current_Stock",
    "Stock_Sharjah", "Stock_Al_Quoz", "Stock_DIP", "Stock_Abu_Dhabi",
    "ABC_Category", "Avg_Daily_Sales", "Baseline_Velocity", "Consistency_Score",
    "Total_Lifetime_Sales", "Lifespan_Days", "Last_Sold_Date", "Last_Updated_Date",
    "Velocity_Al_Quoz", "Velocity_Sharjah", "Velocity_DIP", "Velocity_Abu_Dhabi"
]

if st.session_state.df_stock_live is None:
    if sheet_payload[3]: st.session_state.df_stock_live = pd.DataFrame(sheet_payload[3])
    else: st.session_state.df_stock_live = pd.DataFrame(columns=TARGET_STOCK_COLS)

df_stock = st.session_state.df_stock_live
df_log = pd.DataFrame(sheet_payload[4]) if sheet_payload[4] else pd.DataFrame(columns=["Date", "Item_Code", "Item_Name", "Transaction_Type", "Qty_Delta", "Voucher_Reference", "Timestamp", "Branch", "Voucher abbreviation"])
df_batches = pd.DataFrame(sheet_payload[5]) if sheet_payload[5] else pd.DataFrame(columns=["Batch_ID", "Upload_Type", "Timestamp"])

for col in TARGET_STOCK_COLS:
    if col not in df_stock.columns:
        if any(k in col for k in ["Velocity", "Stock", "Sales", "Score", "Days"]): df_stock[col] = 0.0
        else: df_stock[col] = ""

for d in [df_stock, df_log, df_batches]:
    if not d.empty:
        for col in d.columns:
            if d[col].dtype == 'object': d[col] = d[col].astype(str).str.strip()

def auto_detect_category(item_name):
    name_upper = str(item_name).upper()
    if "ABS SHEET" in name_upper: return "ABS Sheet"
    if "ACRYLIC ROD" in name_upper: return "Acrylic Rod"
    if "ACRYLIC TUBE" in name_upper: return "Acrylic Tube"
    if "ACRYLIC SHEET" in name_upper: return "Acrylic Sheet"
    if "COLD LAMINATION" in name_upper: return "Cold Lamination"
    if "FOAM BOARD" in name_upper: return "Foam Board"
    if "FREE FOAM SHEET" in name_upper: return "Free Foam Sheet"
    if "HDPE ROD" in name_upper: return "HDPE Rod"
    if "PC TWIN SHEET" in name_upper or "PC TWINSHEET" in name_upper: return "PC Twin Sheet"
    if "PC ROLL" in name_upper: return "PC Roll"
    if "PC SHEET" in name_upper: return "PC Sheet"
    if "PVC SHEET" in name_upper: return "PVC Sheet"
    if "TEFFLON SHEET" in name_upper or "TEFLON SHEET" in name_upper: return "Tefflon Sheet"
    if "ACP" in name_upper: return "ACP"
    if "GLUE" in name_upper: return "Glue"
    if "TAPE" in name_upper: return "Tape"
    return "Uncategorized"

def process_daily_sales_intelligence(stock_df, df_sales_raw, file_date_str, alpha=0.08):
    updated_stock = stock_df.copy()
    
    num_cols = ["Current_Stock", "Avg_Daily_Sales", "Baseline_Velocity", "Consistency_Score", 
                "Total_Lifetime_Sales", "Lifespan_Days", "Velocity_Al_Quoz", "Velocity_Sharjah", 
                "Velocity_DIP", "Velocity_Abu_Dhabi", "Stock_Sharjah", "Stock_Al_Quoz", "Stock_DIP", "Stock_Abu_Dhabi"]
    for c in num_cols:
        if c not in updated_stock.columns: updated_stock[c] = 0.0
        updated_stock[c] = pd.to_numeric(updated_stock[c], errors='coerce').fillna(0.0)

    try: current_file_date = datetime.strptime(file_date_str, "%Y-%m-%d").date()
    except Exception: current_file_date = datetime.now().date()

    updated_stock["Item_Code"] = updated_stock["Item_Code"].apply(clean_item_code)

    sales_summary = {}
    branch_summary = {}
    
    for _, r in df_sales_raw.iterrows():
        icode = clean_item_code(r["Item_Code"])
        qty = float(r["Qty_Sold"])
        vtype = str(r.get("Voucher_Type", "")).upper()
        branch = str(r.get("Branch", "")).strip()
        
        net_q = -abs(qty) if "SRTS" in vtype else abs(qty)
        sales_summary[icode] = sales_summary.get(icode, 0.0) + net_q
        
        if branch:
            if icode not in branch_summary: branch_summary[icode] = {}
            branch_summary[icode][branch] = branch_summary[icode].get(branch, 0.0) + net_q

    for idx, row in updated_stock.iterrows():
        sku = clean_item_code(row["Item_Code"])
        last_up = str(row.get("Last_Updated_Date", "")).strip()
        
        if last_up and last_up.lower() not in ["nan", "none", ""]:
            try:
                last_dt = datetime.strptime(last_up, "%Y-%m-%d").date()
                delta_days = max(1, (current_file_date - last_dt).days)
            except Exception: delta_days = 1
        else: delta_days = 1

        old_vel = float(row["Baseline_Velocity"])
        old_score = float(row["Consistency_Score"])
        
        if sku in sales_summary:
            net_daily_qty = sales_summary[sku]
            spike_cap = max(old_vel * 3.0, 10.0)
            regular_qty = min(net_daily_qty, spike_cap) if net_daily_qty > 0 else net_daily_qty
            
            effective_rate = regular_qty / delta_days
            new_vel = (alpha * effective_rate) + ((1.0 - alpha) * old_vel)
            new_score = (alpha * 1.0) + ((1.0 - alpha) * old_score)
            
            updated_stock.at[idx, "Total_Lifetime_Sales"] += net_daily_qty
            updated_stock.at[idx, "Current_Stock"] -= net_daily_qty
            if net_daily_qty > 0: updated_stock.at[idx, "Last_Sold_Date"] = file_date_str
        else:
            decay = (1.0 - alpha) ** delta_days
            new_vel = old_vel * decay
            new_score = old_score * decay

        if sku in branch_summary:
            b_map = {
                "Dubai": "Velocity_Al_Quoz", "Al Quoz": "Velocity_Al_Quoz",
                "Sharjah": "Velocity_Sharjah", "DIP": "Velocity_DIP",
                "Abu Dhabi": "Velocity_Abu_Dhabi"
            }
            for b_key, b_col in b_map.items():
                b_qty = sum(q for b_name, q in branch_summary[sku].items() if b_key.lower() in b_name.lower())
                if b_qty != 0:
                    old_b_vel = float(row[b_col])
                    eff_b_rate = min(b_qty, max(old_b_vel * 3.0, 10.0)) / delta_days
                    updated_stock.at[idx, b_col] = round((alpha * eff_b_rate) + ((1.0 - alpha) * old_b_vel), 4)

        updated_stock.at[idx, "Baseline_Velocity"] = round(max(0.0, new_vel), 4)
        updated_stock.at[idx, "Avg_Daily_Sales"] = round(max(0.0, new_vel), 2)
        updated_stock.at[idx, "Consistency_Score"] = round(max(0.0, min(1.0, new_score)), 4)
        updated_stock.at[idx, "Lifespan_Days"] += delta_days
        updated_stock.at[idx, "Last_Updated_Date"] = file_date_str

    def categorize_sku(r):
        vel = float(r["Baseline_Velocity"])
        score = float(r["Consistency_Score"])
        
        if vel >= 2.0 and score >= 0.35: return "Class A (Fast & Consistent)"
        elif vel >= 0.5 or (r["Total_Lifetime_Sales"] > 100 and score < 0.20): return "Class B (Erratic / Bulk-Driven)"
        elif vel > 0.02: return "Class C (Slow Moving)"
        else: return "Class D (Dead Stock / Inactive)"

    updated_stock["ABC_Category"] = updated_stock.apply(categorize_sku, axis=1)
    return updated_stock

# =====================================================
# INTELLIGENT ALERT DRAFTER (STRICT DATE NORMALIZATION)
# =====================================================
def evaluate_and_queue_branch_alerts(updated_stock, fresh_sh, target_branch, recipient_phone, current_log_df):
    """
    Evaluates stockout risks, strictly parses mixed Google Sheet dates to 
    enforce the 30-day cooldown, and normalizes timestamps so MRNs correctly
    break the cycle only on subsequent days.
    """
    if not fresh_sh: return

    branch_map = {
        "Sharjah": ("Stock_Sharjah", "Velocity_Sharjah"),
        "Al Quoz": ("Stock_Al_Quoz", "Velocity_Al_Quoz"),
        "DIP": ("Stock_DIP", "Velocity_DIP"),
        "Abu Dhabi": ("Stock_Abu_Dhabi", "Velocity_Abu_Dhabi")
    }
    
    target_stock_col, target_vel_col = branch_map[target_branch]
    donor_branches = [(v[0], k) for k, v in branch_map.items() if k != target_branch]

    try:
        try:
            ws_queue = fresh_sh.worksheet("WhatsApp_Alert_Queue")
            df_queue = pd.DataFrame(ws_queue.get_all_records())
        except Exception:
            ws_queue = fresh_sh.add_worksheet(title="WhatsApp_Alert_Queue", rows=100, cols=5)
            ws_queue.append_row(["Date", "SKU", "Recipient", "Message", "Status"])
            df_queue = pd.DataFrame()

        recently_alerted_skus = {}
        active_skus = set()

        if not df_queue.empty and "SKU" in df_queue.columns and "Date" in df_queue.columns:
            # Filter queue for this specific branch
            branch_alerts = df_queue[df_queue["Message"].astype(str).str.contains(target_branch, case=False, na=False)].copy()

            # Safely parse dates, accommodating YYYY-MM-DD or DD/MM/YYYY formatting by Sheets
            branch_alerts["Date_Parsed"] = pd.to_datetime(branch_alerts["Date"], format="%Y-%m-%d", errors='coerce')
            branch_alerts["Date_Parsed"] = branch_alerts["Date_Parsed"].fillna(pd.to_datetime(branch_alerts["Date"], errors='coerce', dayfirst=True))
            # Normalize to 00:00:00 midnight for clean calendar-day comparison
            branch_alerts["Date_Parsed"] = branch_alerts["Date_Parsed"].dt.normalize()

            # Identify alerts sitting in Draft/Pending to prevent double queueing
            if "Status" in branch_alerts.columns:
                active_alerts = branch_alerts[branch_alerts["Status"].astype(str).str.upper().isin(["DRAFT", "PENDING"])]
                active_skus = set(active_alerts["SKU"].astype(str).apply(clean_item_code).str.upper())

            # Find the most recent alert date per SKU
            for _, r in branch_alerts.iterrows():
                s = clean_item_code(str(r["SKU"])).upper()
                d = r["Date_Parsed"]
                if pd.notnull(d):
                    if s not in recently_alerted_skus or d > recently_alerted_skus[s]:
                        recently_alerted_skus[s] = d

        # Prepare log data to detect MRNs (Replenishment Breaker)
        if not current_log_df.empty:
            df_log_copy = current_log_df.copy()
            # Normalize log dates so same-day MRNs don't falsely evaluate as ">" the alert
            df_log_copy["Date_Parsed"] = pd.to_datetime(df_log_copy.get("Timestamp", df_log_copy.get("Date")), errors='coerce').dt.normalize()
            df_log_copy["Qty_Delta"] = pd.to_numeric(df_log_copy["Qty_Delta"], errors='coerce').fillna(0.0)
        else:
            df_log_copy = pd.DataFrame()

        alerts_to_send = []
        today_str = datetime.now().strftime("%Y-%m-%d")
        today_normalized = pd.Timestamp(datetime.now().date())

        for _, row in updated_stock.iterrows():
            sku = clean_item_code(row.get("Item_Code", "")).upper()
            iname = str(row.get("Item_Name", "")).strip()

            if not sku or sku in ["NAN", "NONE", ""]:
                continue

            # 1. Block: Ignore if currently pending or draft
            if sku in active_skus:
                continue

            # 2. Block: Enforce Cooldown UNLESS MRN replenished it strictly AFTER the alert day
            if sku in recently_alerted_skus:
                last_alert_date = recently_alerted_skus[sku]
                days_since = (today_normalized - last_alert_date).days

                if days_since < 30:
                    cooldown_broken = False
                    
                    if not df_log_copy.empty:
                        # Check if a positive inbound (MRN) arrived strictly AFTER the last alert date
                        recent_inbounds = df_log_copy[
                            (df_log_copy["Item_Code"].astype(str).apply(clean_item_code).str.upper() == sku) &
                            (df_log_copy["Date_Parsed"] > last_alert_date) &
                            (df_log_copy["Qty_Delta"] > 0)
                        ]
                        if not recent_inbounds.empty:
                            cooldown_broken = True  # It was restocked and now emptied again!
                            
                    if not cooldown_broken:
                        continue  # Cooldown holds, skip

            b_stock = float(row.get(target_stock_col, 0.0))
            b_vel = float(row.get(target_vel_col, 0.0))
            b_runway = (b_stock / b_vel) if b_vel > 0 else 999.0

            if b_vel > 0.05 and (b_stock <= 5 or b_runway <= 7.0):
                eligible_donors = []
                for donor_col, branch_name in donor_branches:
                    d_qty = float(row.get(donor_col, 0.0))
                    if d_qty > 30: 
                        eligible_donors.append(f"{branch_name} ({int(d_qty)} pcs)")

                if eligible_donors:
                    donor_text = ", ".join(eligible_donors)
                    message = (
                        f"⚠️ *{target_branch.upper()} STOCKOUT ALERT* ⚠️\n"
                        f"----------------------------------------------------------------\n"
                        f"• *SKU:* `{sku}` — {iname}\n"
                        f"• *{target_branch} Balance:* {int(b_stock)} units ({b_runway:.1f} days runway left)\n"
                        f"• *{target_branch} Run-Rate:* {b_vel:.2f} units/day\n"
                        f"• *Surplus Available:* {donor_text}\n"
                        f"👉 *Action Required:* Request immediate internal stock transfer."
                    )
                    alerts_to_send.append([today_str, sku, recipient_phone, message, "Draft"])

        if alerts_to_send:
            ws_queue.append_rows(alerts_to_send)
            st.success(f"📝 Generated {len(alerts_to_send)} new alert draft(s) for {target_branch}.")
        else:
            st.info(f"ℹ️ No new {target_branch} stockout alerts required (items stable, cooling down, or already queued).")

    except Exception as e:
        st.error(f"Failed to scan and draft {target_branch} WhatsApp alerts: {e}")

st.sidebar.markdown("### ⚙️ INVENTORY FILTER")
if not df_stock.empty:
    df_stock["Product_Category"] = df_stock["Product_Category"].replace("", "Uncategorized").fillna("Uncategorized")
    cat_options = ["All Categories"] + sorted(df_stock["Product_Category"].unique().tolist())
else: cat_options = ["All Categories"]

selected_category_filter = st.sidebar.selectbox("Filter by Material Group", cat_options)
item_search = st.sidebar.text_input("🔍 Search Item Code / Description")

if not df_stock.empty:
    df_stock["Current_Stock"] = pd.to_numeric(df_stock["Current_Stock"], errors='coerce').fillna(0)
    df_stock["Avg_Daily_Sales"] = pd.to_numeric(df_stock["Avg_Daily_Sales"], errors='coerce').fillna(0.0)
    df_stock["Days_of_Coverage"] = df_stock.apply(
        lambda r: 999 if r["Avg_Daily_Sales"] <= 0 else round(r["Current_Stock"] / r["Avg_Daily_Sales"], 1), axis=1
    )

filt_stock = df_stock.copy()
if not filt_stock.empty and selected_category_filter != "All Categories":
    filt_stock = filt_stock[filt_stock["Product_Category"] == selected_category_filter]

st.markdown(f"### 📊 Inventory Summary: {selected_category_filter}")

total_skus = len(filt_stock) if not filt_stock.empty else 0
if not filt_stock.empty:
    stockout_count = len(filt_stock[(filt_stock["ABC_Category"].str.contains("Class A", na=False)) & (filt_stock["Days_of_Coverage"] <= 7)])
    a_count = len(filt_stock[filt_stock["ABC_Category"].str.contains("Class A", na=False)])
    dead_stock_count = len(filt_stock[filt_stock["ABC_Category"].str.contains("Class D", na=False)])
else: stockout_count = dead_stock_count = a_count = 0

kpi1, kpi2, kpi3, kpi4 = st.columns(4)
kpi1.metric("SKU COUNT IN FILTER", total_skus)
kpi2.metric("FAST MOVING (CLASS A)", a_count)
kpi3.metric("STOCKOUT RISK (<7 DAYS)", stockout_count, delta="Check Class A", delta_color="inverse")
kpi4.metric("DEAD STOCK (CLASS D)", dead_stock_count, delta="Check Inactive", delta_color="off")

st.markdown(" ")
segment_view = st.radio(
    "Filter Grid Segment View:",
    options=["📋 Show All Rows", "🟩 Fast Moving (Class A)", "🟨 Erratic / Bulk (Class B)", "🚨 High Risk (<7 Days Coverage)", "📉 Dead Stock (Class D)"],
    horizontal=True
)

if segment_view == "🟩 Fast Moving (Class A)":
    filt_stock = filt_stock[filt_stock["ABC_Category"].str.contains("Class A", na=False)]
elif segment_view == "🟨 Erratic / Bulk (Class B)":
    filt_stock = filt_stock[filt_stock["ABC_Category"].str.contains("Class B", na=False)]
elif segment_view == "🚨 High Risk (<7 Days Coverage)":
    filt_stock = filt_stock[(filt_stock["ABC_Category"].str.contains("Class A", na=False)) & (filt_stock["Days_of_Coverage"] <= 7)]
elif segment_view == "📉 Dead Stock (Class D)":
    filt_stock = filt_stock[filt_stock["ABC_Category"].str.contains("Class D", na=False)]

if item_search:
    filt_stock = filt_stock[
        filt_stock["Item_Code"].astype(str).str.contains(item_search, case=False, na=False) | 
        filt_stock["Item_Name"].astype(str).str.contains(item_search, case=False, na=False)
    ]

st.markdown("---")

if not is_admin:
    st.info("🔒 Stock adjustments and data ingestion engines locked. Displaying running terminal logs in read-only mode.")
else:
    col_left, col_right = st.columns(2)
    
    with col_left:
        st.markdown("<div class='upload-box'><h5>📥 Process Incoming Stock (MRN / Opening Balance)</h5>", unsafe_allow_html=True)
        mrn_file = st.file_uploader("Upload Inbound Excel (.xlsx)", type=["xlsx"], key="mrn_loader")
        if mrn_file:
            df_mrn_raw = pd.read_excel(mrn_file, engine="openpyxl")
            df_mrn_raw.columns = [str(c).strip() for c in df_mrn_raw.columns]
            req_mrn = ["Date", "Document No.", "Item.Code", "Item.Name", "Quantity"]
            if all(c in df_mrn_raw.columns for c in req_mrn):
                unique_batch = str(df_mrn_raw["Document No."].iloc[0]).strip()
                if not df_batches.empty and unique_batch in df_batches["Batch_ID"].values:
                    st.error(f"🛑 Double-Upload Blocked! Document Batch `{unique_batch}` has already been processed.")
                else:
                    if st.button("⚡ INTEGRATE INBOUND LOG INTO STOCK"):
                        fresh_sh = get_fresh_google_sheet_file()
                        if not fresh_sh: st.error("🚨 Cloud Write Connection Failed. Please try again.")
                        else:
                            fresh_ws_stock = fresh_sh.get_worksheet(3)
                            try: fresh_ws_log = fresh_sh.worksheet("Daily_Snapshot_Log")
                            except Exception: fresh_ws_log = fresh_sh.get_worksheet(4)
                            fresh_ws_batches = fresh_sh.get_worksheet(5)

                            timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            new_logs, new_stock_map = [], {}
                            for _, row in df_mrn_raw.iterrows():
                                icode = clean_item_code(row["Item.Code"])
                                iname = str(row["Item.Name"]).strip()
                                qty = float(row["Quantity"])
                                new_logs.append([str(row["Date"]), icode, iname, "MRN", qty, unique_batch, timestamp_str, "Central Log", "MRN"])
                                new_stock_map[icode] = {"Item_Name": iname, "Qty": qty}
                            
                            updated_stock = df_stock.copy()
                            for code, info in new_stock_map.items():
                                if not updated_stock.empty and code in updated_stock["Item_Code"].values:
                                    updated_stock.loc[updated_stock["Item_Code"] == code, "Current_Stock"] += info["Qty"]
                                else:
                                    guessed_cat = auto_detect_category(info["Item_Name"])
                                    new_rows_data = {
                                        "Item_Code": code, "Item_Name": info["Item_Name"], "Product_Category": guessed_cat, 
                                        "Current_Stock": info["Qty"], "ABC_Category": "Class C (Slow Moving)", "Avg_Daily_Sales": 0.0, 
                                        "Baseline_Velocity": 0.0, "Consistency_Score": 0.0, "Total_Lifetime_Sales": 0.0, 
                                        "Lifespan_Days": 1.0, "Last_Sold_Date": "", "Last_Updated_Date": datetime.now().strftime("%Y-%m-%d")
                                    }
                                    for b_col in ["Velocity_Al_Quoz", "Velocity_Sharjah", "Velocity_DIP", "Velocity_Abu_Dhabi"]:
                                        new_rows_data[b_col] = 0.0
                                    updated_stock = pd.concat([updated_stock, pd.DataFrame([new_rows_data])], ignore_index=True)
                                    
                            if fresh_ws_stock:
                                fresh_ws_stock.clear()
                                fresh_ws_stock.append_rows([TARGET_STOCK_COLS] + updated_stock[TARGET_STOCK_COLS].fillna("").astype(str).values.tolist())
                            if fresh_ws_log: fresh_ws_log.append_rows(new_logs)
                            if fresh_ws_batches: fresh_ws_batches.append_rows([[unique_batch, "MRN", timestamp_str]])
                            
                            st.session_state.df_stock_live = updated_stock
                            st.success(f"Inbound Sheet Data `{unique_batch}` incorporated successfully!")
                            st.rerun()
            else: st.error(f"Missing column fields. File must match format: {req_mrn}")
        st.markdown("</div>", unsafe_allow_html=True)

    with col_right:
        st.markdown("<div class='upload-box'><h5>📤 Process Outgoing Sales (Daily Ledger)</h5>", unsafe_allow_html=True)
        sales_file = st.file_uploader("Upload Daily Sales Sheet (.xlsx)", type=["xlsx"], key="sales_loader")
        if sales_file:
            df_sales_raw = pd.read_excel(sales_file, engine="openpyxl")
            df_sales_raw.columns = [str(c).strip() for c in df_sales_raw.columns]
            cols = df_sales_raw.columns.tolist()
            def find_col(guesses):
                for g in guesses:
                    for c in cols:
                        if g.lower() in c.lower(): return c
                return cols[0] if cols else ""
                
            match_date = st.selectbox("Match Sales [Date]:", cols, index=cols.index(find_col(["date", "posting"])))
            match_vouch = st.selectbox("Match Sales [Voucher No]:", cols, index=cols.index(find_col(["document", "voucher", "invoice"])))
            match_code = st.selectbox("Match Sales [Item Code]:", cols, index=cols.index(find_col(["item.code", "item_code", "code"])))
            match_name = st.selectbox("Match Sales [Item Name]:", cols, index=cols.index(find_col(["item.name", "item_name", "description"])))
            match_qty = st.selectbox("Match Sales [Quantity]:", cols, index=cols.index(find_col(["quantity", "qty", "sold"])))
            match_branch = st.selectbox("Match Sales [Branch]:", cols, index=cols.index(find_col(["branch", "location", "warehouse"])))
            match_voucher_type = st.selectbox("Match Sales [Voucher Abbreviation]:", cols, index=cols.index(find_col(["abbreviation", "type", "voucher type", "voucher abbreviation"])))
            
            sales_batch_id = str(df_sales_raw[match_vouch].iloc[0]).strip() + "_SALES"
            if not df_batches.empty and sales_batch_id in df_batches["Batch_ID"].values:
                st.error("🛑 Double-Upload Blocked! This daily sales spreadsheet has already been deducted from inventory.")
            else:
                if st.button("⚡ EXECUTE STATEFUL DEMAND ENGINE"):
                    fresh_sh = get_fresh_google_sheet_file()
                    if not fresh_sh: 
                        st.error("🚨 Cloud Write Connection Failed. Please try again.")
                    else:
                        fresh_ws_stock = fresh_sh.get_worksheet(3)
                        try: fresh_ws_log = fresh_sh.worksheet("Daily_Snapshot_Log")
                        except Exception: fresh_ws_log = fresh_sh.get_worksheet(4)
                        fresh_ws_batches = fresh_sh.get_worksheet(5)

                        df_sales_raw["Parsed_Date"] = pd.to_datetime(df_sales_raw[match_date], errors='coerce').dt.strftime('%Y-%m-%d')
                        df_sales_raw = df_sales_raw.dropna(subset=["Parsed_Date"]).sort_values("Parsed_Date")
                        
                        unique_dates = sorted(df_sales_raw["Parsed_Date"].unique())
                        
                        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        updated_stock = df_stock.copy()

                        updated_stock["Item_Code"] = updated_stock["Item_Code"].apply(clean_item_code)

                        progress_bar = st.progress(0)
                        st.info(f"⏳ Processing history across {len(unique_dates)} distinct date(s)...")

                        for idx_d, d_str in enumerate(unique_dates):
                            day_sales = df_sales_raw[df_sales_raw["Parsed_Date"] == d_str]
                            
                            clean_sales_records = []
                            for _, row in day_sales.iterrows():
                                clean_code = clean_item_code(row[match_code])
                                clean_sales_records.append({
                                    "Item_Code": clean_code,
                                    "Item_Name": str(row[match_name]).strip(),
                                    "Qty_Sold": float(row[match_qty]),
                                    "Branch": str(row[match_branch]).strip(),
                                    "Voucher_Type": str(row[match_voucher_type]).strip()
                                })
                            
                            df_clean_day = pd.DataFrame(clean_sales_records)
                            updated_stock = process_daily_sales_intelligence(updated_stock, df_clean_day, d_str)
                            progress_bar.progress((idx_d + 1) / len(unique_dates))

                        new_log_rows = []
                        for _, row in df_sales_raw.iterrows():
                            clean_code = clean_item_code(row[match_code])
                            new_log_rows.append([
                                str(row[match_date]), clean_code, str(row[match_name]).strip(),
                                "Sales", -abs(float(row[match_qty])), str(row[match_vouch]).strip(),
                                timestamp_str, str(row[match_branch]).strip(), str(row[match_voucher_type]).strip()
                            ])

                        sales_batch_id = f"BOOTSTRAP_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

                        if fresh_ws_stock:
                            fresh_ws_stock.clear()
                            fresh_ws_stock.append_rows([TARGET_STOCK_COLS] + updated_stock[TARGET_STOCK_COLS].fillna("").astype(str).values.tolist())
                        if fresh_ws_log: 
                            fresh_ws_log.append_rows(new_log_rows)
                        if fresh_ws_batches: 
                            fresh_ws_batches.append_rows([[sales_batch_id, "Sales_Bootstrap", timestamp_str]])
                        
                        st.session_state.df_stock_live = updated_stock
                        st.success("Stateful demand intelligence successfully updated day-by-day!")
                        st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

# =====================================================
# DEDICATED STOCK ALERT EVALUATION & APPROVAL GATEWAY
# =====================================================
if is_admin:
    st.markdown("---")
    st.markdown("### 📲 Automated Stockout Alert Managers")
    
    # --- SHARJAH SECTION ---
    st.markdown("#### 📍 Sharjah Operations")
    col_shj, col_shj_info = st.columns([1, 2])
    with col_shj:
        if st.button("🔍 SCAN SHARJAH STOCK & DRAFT ALERTS", use_container_width=True):
            fresh_sh = get_fresh_google_sheet_file()
            if fresh_sh and st.session_state.df_stock_live is not None:
                # Pass df_log into the queue evaluator for MRN detection
                evaluate_and_queue_branch_alerts(st.session_state.df_stock_live, fresh_sh, "Sharjah", "971582577622", df_log)
                st.rerun()
            else: st.error("🚨 Cloud connection failure.")
    with col_shj_info:
        st.caption("Evaluates Sharjah stock against DIP, Al Quoz, and Abu Dhabi surplus.")

    # --- AL QUOZ SECTION ---
    st.markdown("#### 📍 Al Quoz Operations")
    col_aq, col_aq_info = st.columns([1, 2])
    with col_aq:
        if st.button("🔍 SCAN AL QUOZ STOCK & DRAFT ALERTS", use_container_width=True):
            fresh_sh = get_fresh_google_sheet_file()
            if fresh_sh and st.session_state.df_stock_live is not None:
                # Pass df_log into the queue evaluator for MRN detection
                evaluate_and_queue_branch_alerts(st.session_state.df_stock_live, fresh_sh, "Al Quoz", "971555795246", df_log) 
                st.rerun()
            else: st.error("🚨 Cloud connection failure.")
    with col_aq_info:
        st.caption("Evaluates Al Quoz stock against Sharjah, DIP, and Abu Dhabi surplus.")

    # --- UNIFIED MASTER APPROVAL QUEUE ---
    st.markdown("---")
    st.markdown("#### 📋 Pending Alert Approvals (All Branches)")
    fresh_sh = get_fresh_google_sheet_file()
    if fresh_sh:
        try:
            ws_queue = fresh_sh.worksheet("WhatsApp_Alert_Queue")
            df_queue = pd.DataFrame(ws_queue.get_all_records())
            
            if not df_queue.empty and "Status" in df_queue.columns:
                draft_alerts = df_queue[df_queue["Status"].astype(str).str.upper() == "DRAFT"]
                
                if draft_alerts.empty:
                    st.info("✅ No pending alert drafts requiring approval right now.")
                else:
                    st.warning(f"⚠️ There are **{len(draft_alerts)}** alert draft(s) across your network waiting for review.")
                    
                    col_bulk, col_pad = st.columns([1, 2])
                    with col_bulk:
                        if st.button("🚀 BULK APPROVE ALL PENDING DRAFTS", use_container_width=True):
                            df_queue.loc[df_queue["Status"].astype(str).str.upper() == "DRAFT", "Status"] = "Pending"
                            ws_queue.clear()
                            ws_queue.append_rows([df_queue.columns.tolist()] + df_queue.fillna("").astype(str).values.tolist())
                            st.success("✅ All drafts approved and queued for dispatch!")
                            st.rerun()
                            
                    st.markdown(" ") # Spacer
                    
                    for idx, row in draft_alerts.iterrows():
                        sheet_row_num = idx + 2  
                        sku_code = row.get("SKU", "N/A")
                        msg_preview = row.get("Message", "")
                        
                        with st.expander(f"📦 Review Alert Draft for SKU: `{sku_code}`", expanded=False):
                            st.code(msg_preview, language="markdown")
                            
                            col_app, col_rej = st.columns([1, 1])
                            with col_app:
                                if st.button(f"✅ Approve & Queue", key=f"app_{sku_code}_{sheet_row_num}"):
                                    ws_queue.update_cell(sheet_row_num, 5, "Pending")
                                    st.success(f"🚀 Alert for `{sku_code}` approved! Status set to Pending.")
                                    st.rerun()
                            with col_rej:
                                if st.button(f"❌ Reject / Cancel", key=f"rej_{sku_code}_{sheet_row_num}"):
                                    ws_queue.update_cell(sheet_row_num, 5, "Rejected")
                                    st.info(f"🗑️ Alert for `{sku_code}` cancelled.")
                                    st.rerun()
        except Exception as e:
            st.error(f"⚠️ Error loading WhatsApp Alert Queue: {e}")

grid_header_col, download_btn_col = st.columns([3, 1])
with grid_header_col: st.markdown("### 📜 Material Segment Tracking Ledger")

def generate_professional_excel(dataframe, segment_name):
    output = io.BytesIO()
    clean_df = dataframe.copy()
    clean_df["Last_Sold_Date"] = clean_df["Last_Sold_Date"].replace("", "Never Tracked").fillna("Never Tracked")
    
    clean_df = clean_df[[
        "Item_Code", "Item_Name", "Product_Category", "Current_Stock", 
        "ABC_Category", "Avg_Daily_Sales", "Consistency_Score", "Days_of_Coverage", "Last_Sold_Date"
    ]].rename(columns={
        "Item_Code": "Item Code",
        "Item_Name": "Product Specification / Description",
        "Product_Category": "Material Group",
        "Current_Stock": "Current Balance",
        "ABC_Category": "Intelligence Velocity Profile",
        "Avg_Daily_Sales": "Baseline Run-Rate (Units/Day)",
        "Consistency_Score": "Order Consistency Score",
        "Days_of_Coverage": "Estimated Days Runway",
        "Last_Sold_Date": "Last Active Dispatch Date"
    })

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        clean_df.to_excel(writer, sheet_name="Inventory Report", index=False, startrow=4)
        workbook = writer.book
        worksheet = writer.sheets["Inventory Report"]
        worksheet.views.sheetView[0].showGridLines = True
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        font_family = "Segoe UI"
        navy_header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        
        font_title = Font(name=font_family, size=16, bold=True, color="1B2A4A")
        font_subtitle = Font(name=font_family, size=10, italic=True, color="555555")
        font_headers = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_data = Font(name=font_family, size=10)
        
        thin_side = Side(border_style="thin", color="D1D5DB")
        data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        worksheet["A1"] = "SABIN PLASTIC // ENTERPRISE WAREHOUSE LEDGER"
        worksheet["A1"].font = font_title
        worksheet["A2"] = f"Material Segment Slice: {segment_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet["A2"].font = font_subtitle
        
        header_row = 5
        worksheet.row_dimensions[header_row].height = 26
        for col_idx in range(1, len(clean_df.columns) + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = navy_header_fill
            cell.font = font_headers
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
            cell.border = data_border
            
        for row_idx in range(6, len(clean_df) + 6):
            worksheet.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(clean_df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = data_border
                
                if col_idx in [4, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if col_idx == 4: cell.number_format = '#,##0'
                    if col_idx in [6, 7]: cell.number_format = '#,##0.00'
                    if col_idx == 8: cell.number_format = '#,##0'
                elif col_idx in [1, 3, 5, 9]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    return output.getvalue()

with download_btn_col:
    if not filt_stock.empty:
        excel_data = generate_professional_excel(filt_stock, segment_view)
        st.download_button(
            label="📥 Download Excel Ledger",
            data=excel_data,
            file_name=f"Sabin_Inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if filt_stock.empty:
    st.info("📌 No items found matching the selected segment filter criteria.")
else:
    df_stock_display = filt_stock.copy()
    df_stock_display["Last_Sold_Date"] = df_stock_display["Last_Sold_Date"].replace("", "Never Tracked").fillna("Never Tracked")
    
    display_columns = [
        "Item_Code", "Item_Name", "Product_Category", "Current_Stock", 
        "ABC_Category", "Avg_Daily_Sales", "Consistency_Score", "Days_of_Coverage", "Last_Sold_Date"
    ]
    
    st.dataframe(
        df_stock_display[display_columns].sort_values(by="Current_Stock", ascending=True),
        hide_index=True,
        column_config={
            "Item_Code": st.column_config.TextColumn("Item Code"),
            "Item_Name": st.column_config.TextColumn("Product Specification / Description"),
            "Product_Category": st.column_config.TextColumn("Material Group"),
            "Current_Stock": st.column_config.NumberColumn("Current Balance", format="%d Units"),
            "ABC_Category": st.column_config.TextColumn("Demand Movement Profile"),
            "Avg_Daily_Sales": st.column_config.NumberColumn("Baseline Run-Rate", format="%.2f Units/Day 📈"),
            "Consistency_Score": st.column_config.NumberColumn("Order Frequency Index", format="%.2f 🎯"),
            "Days_of_Coverage": st.column_config.NumberColumn("Estimated Runway", format="%d Days ⏳"),
            "Last_Sold_Date": st.column_config.TextColumn("Last Dispatched Activity Date")
        }
    )

if is_admin and not df_stock.empty:
    df_stock["Product_Category"] = df_stock["Product_Category"].astype(str).str.strip()
    uncat_items = df_stock[df_stock["Product_Category"].isin(["Uncategorized", "", "None", "nan"])]
    
    if not uncat_items.empty:
        st.markdown("<div class='admin-box'>⚙️ <b>Autonomous Intelligence Gateway: Global Smart Assignment</b>", unsafe_allow_html=True)
        st.info(f"The system has detected **{len(uncat_items)}** unique item(s) currently marked as `Uncategorized` from your uploads.")
        
        known_cats = sorted(list(set(df_stock["Product_Category"].unique()) - {"Uncategorized", "", "None", "nan"}))
        target_row = uncat_items.iloc[0]
        st.warning(f"**Target Code:** `{target_row['Item_Code']}` | **Specification:** `{target_row['Item_Name']}`")
        
        assign_col1, assign_col2 = st.columns(2)
        with assign_col1:
            chosen_existing = st.selectbox("Assign to an Existing Material Group:", ["-- Create Completely New --"] + known_cats)
        with assign_col2:
            custom_new_cat = st.text_input("Or Type a Brand New Category Name (e.g., Mirror Sheet, Rods, Adhesives):")
            
        if st.button("💾 SAVE & RE-INDEX ALL RELATED ITEMS"):
            fresh_sh = get_fresh_google_sheet_file()
            if not fresh_sh: st.error("🚨 Cloud Write Connection Failed. Please try again.")
            else:
                fresh_ws_stock = fresh_sh.get_worksheet(3)
                final_cat_selection = custom_new_cat.strip() if chosen_existing == "-- Create Completely New --" and custom_new_cat.strip() != "" else chosen_existing
                
                if final_cat_selection in ["-- Create Completely New --", ""]:
                    st.error("Please enter or choose a valid target category label before clicking update.")
                else:
                    updated_stock = df_stock.copy()
                    target_code = clean_item_code(target_row['Item_Code'])
                    target_description = str(target_row['Item_Name']).upper().strip()
                    potential_kw = final_cat_selection.upper().replace("SHEET", "").replace("ROD", "").strip()
                    
                    if len(potential_kw) > 2 and potential_kw in target_description: matched_keyword = potential_kw
                    else:
                        words = [w for w in target_description.split() if len(w) > 2]
                        matched_keyword = words[0] if words else ""

                    updated_stock["Item_Code_Str"] = updated_stock["Item_Code"].apply(clean_item_code)
                    if matched_keyword != "":
                        updated_stock["Item_Name_Upper"] = updated_stock["Item_Name"].astype(str).str.upper()
                        mask = (updated_stock["Product_Category"].isin(["Uncategorized", "", "None", "nan"])) & \
                               (updated_stock["Item_Name_Upper"].str.contains(matched_keyword, na=False))
                        updated_stock.loc[mask, "Product_Category"] = final_cat_selection
                        updated_stock.loc[updated_stock["Item_Code_Str"] == target_code, "Product_Category"] = final_cat_selection
                        updated_stock.drop(columns=["Item_Name_Upper"], inplace=True)
                    else:
                        updated_stock.loc[updated_stock["Item_Code_Str"] == target_code, "Product_Category"] = final_cat_selection
                    
                    updated_stock.drop(columns=["Item_Code_Str"], inplace=True)
                    
                    if fresh_ws_stock:
                        try:
                            fresh_ws_stock.clear()
                            fresh_ws_stock.append_rows([TARGET_STOCK_COLS] + updated_stock[TARGET_STOCK_COLS].fillna("").astype(str).values.tolist())
                            st.session_state.df_stock_live = updated_stock
                            st.success(f"Successfully updated items to '{final_cat_selection}'!")
                            st.rerun()
                        except Exception as cloud_err: st.error(f"Write operation failed via cloud API: {cloud_err}")
        st.markdown("</div>", unsafe_allow_html=True)
# --- Floating Copilot Engine ---
inject_floating_copilot()
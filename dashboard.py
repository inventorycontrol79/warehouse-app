import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # Premium formatting utilities

# 1. Page Configuration & Modern UI Styles
st.set_page_config(
    page_title="Warehouse Verification Control",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Professional UI CSS customization
st.markdown("""
    <style>
    .main .block-container { padding-top: 2rem; padding-bottom: 2rem; }
    h1, h2, h3 { font-family: 'Segoe UI', Helvetica, sans-serif; font-weight: 600; color: #2c3e50; }
    div[data-testid="stMetricValue"] { font-size: 28px !important; font-weight: 700 !important; color: #1e293b !important; }
    div[data-testid="stMetricLabel"] { font-size: 14px !important; font-weight: 500 !important; color: #64748b !important; }
    .custom-hr { border: 0; height: 1px; background: #e2e8f0; margin: 1.5rem 0; }
    </style>
""", unsafe_allow_html=True)

# 2. Perfect Absolute Path to your inventory file
CSV_FILE_PATH = "D:\\Afsal\\Whatsapp_bot\\inventory.csv"

def load_warehouse_data():
    try:
        df = pd.read_csv(CSV_FILE_PATH)
        
        # Clean column string whitespaces to prevent mapping/grouping issues
        df.columns = df.columns.str.strip()
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
            
        return df
    except FileNotFoundError:
        st.error(f"❌ File not found at: {CSV_FILE_PATH}")
        st.info("Please make sure your file is saved precisely as 'inventory.csv' inside your 'D:\\Afsal\\Whatsapp_bot' directory.")
        st.stop()

df_raw = load_warehouse_data()

# Default standard color scheme palette for charts
status_colors = {
    'Dispatched': '#2ecc71',           # Vibrant soft green
    'Pending': '#f1c40f',              # Clear warning yellow
    'Return': '#e74c3c',               # Distinct soft red
    'Returned': '#e74c3c',             # Alternative mapping variant
    'Verification Failed': '#d35400'   # Clean orange fallback
}

# 3. Sidebar Filtering (Control Center)
st.sidebar.title("🛠️ Control Center")
st.sidebar.markdown("Filter active dispatch lists below.")

# Global Text Search Box
search_query = st.sidebar.text_input("🔍 Global Search", "", placeholder="Search DO # or Phone...")

# Warehouse Hub Multi-select filter
all_warehouses = sorted(df_raw['Warehouse_Name'].unique().tolist())
selected_warehouses = st.sidebar.multiselect("Select Warehouses", options=all_warehouses, default=all_warehouses)

# Status Checkbox Filter
all_statuses = sorted(df_raw['Status'].unique().tolist())
selected_statuses = st.sidebar.multiselect("Filter Statuses", options=all_statuses, default=all_statuses)

# --- RESTORED: Date Picker Filter Option ---
if 'Date_Issued' in df_raw.columns:
    # Ensure standard pandas datetime matching rules apply cleanly
    df_raw['Date_Issued_DT'] = pd.to_datetime(df_raw['Date_Issued'], errors='coerce')
    min_date = df_raw['Date_Issued_DT'].min().date() if not df_raw['Date_Issued_DT'].isnull().all() else None
    max_date = df_raw['Date_Issued_DT'].max().date() if not df_raw['Date_Issued_DT'].isnull().all() else None
    
    if min_date and max_date:
        selected_date_range = st.sidebar.date_input("Filter Date Range", value=(min_date, max_date), min_value=min_date, max_value=max_date)
    else:
        selected_date_range = None
else:
    selected_date_range = None

# Apply active sidebar data rules
df_filtered = df_raw[
    (df_raw['Warehouse_Name'].isin(selected_warehouses)) & 
    (df_raw['Status'].isin(selected_statuses))
].copy()

# Apply Date Filtering constraints if a range slice is specified
if selected_date_range and len(selected_date_range) == 2:
    start_date, end_date = selected_date_range
    df_filtered = df_filtered[
        (df_filtered['Date_Issued_DT'].dt.date >= start_date) & 
        (df_filtered['Date_Issued_DT'].dt.date <= end_date)
    ]

if search_query:
    # Drop date objects explicitly during the query text scanning to prevent data type errors
    search_df = df_filtered.drop(columns=['Date_Issued_DT'], errors='ignore')
    mask = search_df.astype(str).apply(lambda x: x.str.contains(search_query, case=False)).any(axis=1)
    df_filtered = df_filtered[mask]

# Clean up helper column before displaying or exporting data logs
if 'Date_Issued_DT' in df_filtered.columns:
    df_filtered = df_filtered.drop(columns=['Date_Issued_DT'])

# 4. Dashboard Headers
st.title("📦 Logistics Verification & Dispatch Control")
st.caption("Real-time operations management for automated delivery cycles across regional distribution centers.")
st.markdown('<div class="custom-hr"></div>', unsafe_allow_html=True)

# 5. Summary Metric Calculations
total_orders = len(df_filtered)
dispatched_count = len(df_filtered[df_filtered['Status'].str.lower() == 'dispatched'])
pending_count = len(df_filtered[df_filtered['Status'].str.lower() == 'pending'])
return_count = len(df_filtered[df_filtered['Status'].str.lower().str.contains('return')])

efficiency_rate = (dispatched_count / total_orders * 100) if total_orders > 0 else 0.0

m1, m2, m3, m4 = st.columns(4)
m1.metric(label="Total Tracked Load", value=total_orders)
m2.metric(label="✅ Dispatched Orders", value=dispatched_count, delta=f"+{dispatched_count} completed")
m3.metric(label="⏳ Pending Lineup", value=pending_count, delta=f"{pending_count} waiting", delta_color="inverse")
m4.metric(label="🔄 Marked Returns", value=return_count, delta=f"{return_count} flagged", delta_color="off" if return_count==0 else "inverse")

st.markdown(f"**Current Operational Automation Efficiency:** `{efficiency_rate:.1f}%`")
st.markdown('<div class="custom-hr"></div>', unsafe_allow_html=True)

# 6. Side-By-Side Interactive Visualizations
col_chart1, col_chart2 = st.columns([1, 2])

with col_chart1:
    # --- RESTORED: Core Interactive Stock/Milestone Pie Chart ---
    st.subheader("Milestone Distribution")
    if total_orders > 0:
        status_counts = df_filtered['Status'].value_counts().reset_index()
        status_counts.columns = ['Status', 'Count']
        fig_pie = px.pie(
            status_counts, values='Count', names='Status', hole=0.4,
            color='Status', color_discrete_map=status_colors
        )
        fig_pie.update_layout(margin=dict(t=20, b=20, l=10, r=10), showlegend=True, height=280)
        st.plotly_chart(fig_pie, use_container_width=True)
    else:
        st.info("No active records to segment.")

with col_chart2:
    st.subheader("Regional Hub Outputs")
    if total_orders > 0:
        fig_bar = px.bar(
            df_filtered, x='Warehouse_Name', color='Status',
            barmode='group', color_discrete_map=status_colors
        )
        fig_bar.update_layout(margin=dict(t=20, b=20, l=10, r=10), height=280, xaxis_title="Warehouse Location", yaxis_title="Orders")
        st.plotly_chart(fig_bar, use_container_width=True)
    else:
        st.info("No data available to map metrics.")

st.markdown('<div class="custom-hr"></div>', unsafe_allow_html=True)

# 7. Sleek Interactive Data Table Queue
st.subheader("📋 Active Operations Log")
st.dataframe(
    df_filtered,
    use_container_width=True,
    column_config={
        "DO_Number": st.column_config.TextColumn("Delivery Order Number"),
        "Last_4": st.column_config.TextColumn("Ending Sequence", help="Last 4 characters of the order number"),
        "Supervisor_Phone": st.column_config.TextColumn("Supervisor Line"),
        "Date_Issued": st.column_config.TextColumn("Date Registered"),
        "Warehouse_Name": st.column_config.TextColumn("Operational Hub")
    },
    hide_index=True
)

# 8. Premium Single-Sheet Excel Spreadsheet Exporter
st.markdown("### 📥 Professional Data Extraction")
st.caption("Convert your live filter queries into a clean, presentation-ready business spreadsheet.")

HEADER_MAPPING = {
    'DO_Number': 'Delivery Order Reference',
    'Last_4': 'Trailing Match Tag',
    'Supervisor_Phone': 'Contact Registry Number',
    'Status': 'Operational Milestone Status',
    'Date_Issued': 'Manifest Creation Date',
    'Warehouse_Name': 'Dispatched Hub Node'
}

def convert_df_to_excel(df_to_export):
    output = BytesIO()
    
    # Isolate columns and apply mapping rules
    columns_to_keep = [col for col in df_to_export.columns if col in HEADER_MAPPING.keys()]
    final_df = df_to_export[columns_to_keep].rename(columns=HEADER_MAPPING)
    
    # Pre-calculate totals for the embedded dashboard summary cards
    t_orders = len(df_to_export)
    t_dispatched = len(df_to_export[df_to_export['Status'].str.lower() == 'dispatched'])
    t_pending = len(df_to_export[df_to_export['Status'].str.lower() == 'pending'])
    t_returns = len(df_to_export[df_to_export['Status'].str.lower().str.contains('return')])
    t_efficiency = (t_dispatched / t_orders) if t_orders > 0 else 0.0

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        SHEET_NAME = 'Delivery Dispatch status'
        
        # Start writing data table down at row 8 to leave room for summary block at the top
        final_df.to_excel(writer, index=False, sheet_name=SHEET_NAME, startrow=7)
        worksheet = writer.sheets[SHEET_NAME]
        
        # High-End Design Style Tokens
        font_title = Font(name='Segoe UI', size=12, bold=True, color='1F3A52')
        font_kpi_label = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        font_kpi_val = Font(name='Segoe UI', size=11, bold=True, color='1F3A52')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Segoe UI', size=10, color='2C3E50')
        
        # Colors
        fill_kpi_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Slate Gray
        fill_kpi_card = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")   # Clean Light Off-White
        header_fill = PatternFill(start_color="1F3A52", end_color="1F3A52", fill_type="solid")     # Premium Midnight Blue
        
        thin_border_side = Side(style='thin', color='D1D5DB')  # Subtle clean gray grid border lines
        grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # --- BUILD EXECUTIVE KPI CARD BLOCK (Rows 1-5) ---
        worksheet['A1'] = "WAREHOUSE DISPATCH PERFORMANCE SUMMARY"
        worksheet['A1'].font = font_title
        
        # Define KPI Layout Grid
        kpis = [
            ("Total Tracked Orders", f"{t_orders} Pcs", "A", "B"),
            ("Dispatched Orders", f"{t_dispatched} Pcs", "C", "D"),
            ("Pending Orders", f"{t_pending} Pcs", "E", "F"),
            ("Marked Returns", f"{t_returns} Pcs", "G", "H")
        ]
        
        for label, val, start_col, end_col in kpis:
            worksheet.merge_cells(f"{start_col}3:{end_col}3")
            worksheet.merge_cells(f"{start_col}4:{end_col}4")
            
            lbl_cell = worksheet[f"{start_col}3"]
            val_cell = worksheet[f"{start_col}4"]
            
            lbl_cell.value = label
            lbl_cell.font = font_kpi_label
            lbl_cell.fill = fill_kpi_header
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            val_cell.value = val
            val_cell.font = font_kpi_val
            val_cell.fill = fill_kpi_card
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for r in [3, 4]:
                worksheet[f"{start_col}{r}"].border = grid_border
                worksheet[f"{end_col}{r}"].border = grid_border

        # Add Efficiency Summary Card line
        worksheet.merge_cells("A5:D5")
        worksheet["A5"] = f"Current Operational Automation Efficiency:  {t_efficiency:.1%}"
        worksheet["A5"].font = Font(name='Segoe UI', size=10, bold=True, color='2C3E50')
        worksheet["A5"].alignment = Alignment(horizontal="left", vertical="center")

        # --- FORMAT MAIN DATA LOG GRID (Row 8 Downward) ---
        worksheet.row_dimensions[8].height = 26
        for cell in worksheet[8]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Define Executive Soft Pastel Background Conditional Fills
        fill_dispatched = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Polished Mint Green
        fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # Warm Soft Yellow
        fill_return = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Subtle Muted Red/Coral

        # Dynamically locate the 'Operational Milestone Status' column coordinate (1-based)
        mapped_status_col_name = HEADER_MAPPING['Status']
        status_col_idx = final_df.columns.get_loc(mapped_status_col_name) + 1 

        # Style data matrix rows sequentially starting right under header at row 9
        for row_idx in range(9, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20  # Give cells breathing space
            status_value = str(worksheet.cell(row=row_idx, column=status_col_idx).value).strip().lower()
            
            # Match status keywords smoothly
            current_fill = None
            if status_value == 'dispatched':
                current_fill = fill_dispatched
            elif status_value == 'pending':
                current_fill = fill_pending
            elif 'return' in status_value:
                current_fill = fill_return

            # Apply fonts, light borders, and matching colors across the row
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = grid_border
                
                # Center align identifiers, trailing codes, dates, and status fields; left align anything else
                if col_idx in [status_col_idx, 1, 2, 5]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                if current_fill:
                    cell.fill = current_fill
        
        # Prevent layout crunching (###) by expanding column boundaries dynamically
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 5, 16)
            
    return output.getvalue()

if not df_filtered.empty:
    excel_data = convert_df_to_excel(df_filtered)
    st.download_button(
        label="📥 Download Excel Dispatch Manifest",
        data=excel_data,
        file_name="Warehouse_Manifest_Extract.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.download_button("📥 Download Excel Dispatch Manifest", data=b"", disabled=True)